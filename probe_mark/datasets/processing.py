import os

import cv2
import numpy as np
import openpyxl


class ImageExtractor:
    def __init__(self, product_dir: str, save_dir: str):
        self.product_dir = product_dir
        self.save_dir = save_dir

        self.product_id = os.path.basename(product_dir)  # 產品名稱
        self.excel_path = os.path.join(product_dir, "picture.xlsx")  # Excel路徑
        # TODO(hcchen): 暫定產品名稱為保存目錄
        self.save_dir = os.path.join(save_dir, self.product_id)  # 設定保存目錄

        # 加載 Excel 工作簿
        self.workbook = openpyxl.load_workbook(self.excel_path, data_only=True)

        data = self.process_data()
        self.save_data(data)

    def process_data(self) -> dict:
        """處理數據，提取影像和標籤，並進行清理"""
        data = {}
        data = self.extract_images(data)
        data = self.extract_labels(data)
        data = self.remove_problematic(data)
        data = self.remove_incomplete(data)
        data = self.create_masks(data)
        return data

    def extract_images(self, data: dict) -> dict:
        """從工作表中提取影像數據並存入字典"""
        sheet = self.workbook["主要"]
        for img in sheet._images:
            row = img.anchor._from.row + 1
            name = sheet.cell(row, 1).value
            if name not in data:
                data[name] = {}
            data[name]["image"] = self.decode_image(img._data())
        return data

    def extract_labels(self, data: dict) -> dict:
        """從工作表中提取標籤數據並存入字典"""
        sheet = self.workbook["體積面積量測"]
        for img in sheet._images:
            row = img.anchor._from.row + 1
            name = sheet.cell(row, 1).value
            if name not in data:
                data[name] = {}
            data[name]["label"] = self.decode_image(img._data())
        return data

    @staticmethod
    def decode_image(data: bytes) -> np.ndarray:
        """將影像數據從 bytes 解碼為 numpy 陣列"""
        img_array = np.asarray(bytearray(data), dtype=np.uint8)
        return cv2.imdecode(img_array, cv2.IMREAD_COLOR)

    @staticmethod
    def remove_incomplete(data: dict) -> dict:
        """移除不完整的數據條目"""
        incomplete = [
            name
            for name, items in data.items()
            if "image" not in items or "label" not in items
        ]
        for name in incomplete:
            print(f"Missing data for {name}, deleting entry.")
            del data[name]
        return data

    @staticmethod
    def create_masks(data: dict) -> dict:
        """找到圖像與標籤之間的差異，並創建遮罩"""
        for name, items in data.items():
            image, label = items["image"], items["label"]
            diff = cv2.absdiff(image, label)
            gray_diff = cv2.cvtColor(diff, cv2.COLOR_BGR2GRAY)
            items["mask"] = cv2.threshold(gray_diff, 1, 1, cv2.THRESH_BINARY)[1]
            items["mask_view"] = cv2.threshold(gray_diff, 1, 255, cv2.THRESH_BINARY)[1]
        return data

    def remove_problematic(self, data: dict) -> dict:
        """移除特定產品中問題數據條目"""
        if self.product_id == "Artificial_VM_D80_MW120F-HS":
            problematic = ["125-OD100-1", "125-OD100-2", "125-OD100-3"]
            for item in problematic:
                data.pop(item, None)
        return data

    def save_data(self, data: dict):
        """保存處理後的數據至指定目錄"""
        os.makedirs(self.save_dir, exist_ok=True)
        for name, items in data.items():
            dir_path = os.path.join(self.save_dir, name)
            os.makedirs(dir_path, exist_ok=True)
            for tag, image in items.items():
                path = os.path.join(dir_path, f"{tag}.png")
                cv2.imwrite(path, image)


if __name__ == "__main__":
    raw_dir = "data/raw"
    processed_dir = "data/processed"

    for subfolder in os.listdir(raw_dir):
        product_dir = os.path.join(raw_dir, subfolder)
        if os.path.isdir(product_dir):  # 確認該路徑是否為目錄
            extractor = ImageExtractor(product_dir=product_dir, save_dir=processed_dir)
